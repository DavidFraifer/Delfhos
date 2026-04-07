"""
Files Tool - Intelligent file handling for agents
Provides automatic parsing, analysis, and manipulation of various file types
"""

import json
import csv
import io
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import base64
from delfhos.errors import OptionalDependencyError, ToolExecutionError

from ..utils.console import console


class FileHandler:
    """Intelligent file handler with automatic type detection and parsing"""
    
    def __init__(self, file_path: str, task_id: str = None, agent_id: str = None):
        self.file_path = Path(file_path)
        self.task_id = task_id
        self.agent_id = agent_id
        self.metadata = {}
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        self._detect_type()
    
    def _detect_type(self):
        """Detect file type from extension and content"""
        suffix = self.file_path.suffix.lower()
        
        # Map extensions to types
        type_map = {
            '.csv': 'csv',
            '.json': 'json',
            '.txt': 'text',
            '.md': 'markdown',
            '.pdf': 'pdf',
            '.xlsx': 'excel',
            '.xls': 'excel',
            '.jpg': 'image',
            '.jpeg': 'image',
            '.png': 'image',
            '.gif': 'image',
            '.webp': 'image',
            '.svg': 'image',
            '.zip': 'archive',
            '.tar': 'archive',
            '.gz': 'archive',
        }
        
        self.file_type = type_map.get(suffix, 'file')
        self.metadata['extension'] = suffix
        self.metadata['size_bytes'] = self.file_path.stat().st_size
    
    def read_text(self, encoding: str = 'utf-8') -> str:
        """Read file as plain text"""
        try:
            return self.file_path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            # Try with different encodings
            for enc in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    return self.file_path.read_text(encoding=enc)
                except (UnicodeDecodeError, LookupError):
                    continue
            raise ToolExecutionError(tool_name="files", detail="Could not decode file with common encodings")
    
    def read_csv(self, has_header: bool = True) -> List[Dict[str, Any]]:
        """Read CSV file as list of dictionaries"""
        content = self.read_text()
        reader = csv.DictReader(io.StringIO(content)) if has_header else csv.reader(io.StringIO(content))
        
        if has_header:
            return list(reader)
        else:
            # Convert to list of lists
            return [list(row) for row in reader]
    
    def read_json(self) -> Union[Dict, List]:
        """Read JSON file and parse"""
        content = self.read_text()
        return json.loads(content)
    
    def read_excel(self, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """Read Excel file as list of dictionaries"""
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            # Read data rows
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                data.append(row_dict)
            
            return data
        except ImportError:
            raise OptionalDependencyError(
                package="openpyxl",
                detail="Required for reading Excel files (.xlsx/.xls)."
            )
    
    def read_image_base64(self) -> str:
        """Read image as base64 string"""
        with open(self.file_path, 'rb') as f:
            return base64.b64encode(f.read()).decode('utf-8')
    
    def read_pdf_text(self) -> str:
        """Extract text from PDF"""
        try:
            import PyPDF2
            with open(self.file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = []
                for page in reader.pages:
                    text.append(page.extract_text())
                return '\n\n'.join(text)
        except ImportError:
            raise OptionalDependencyError(
                package="PyPDF2",
                detail="Required for extracting text from PDF files."
            )
    
    def to_dict(self) -> Dict[str, Any]:
        """Get file information as dictionary"""
        return {
            'filename': self.file_path.name,
            'file_type': self.file_type,
            'path': str(self.file_path),
            'size_bytes': self.metadata.get('size_bytes', 0),
            'extension': self.metadata.get('extension', ''),
        }


async def load_task_files(task_id: str) -> List[FileHandler]:
    """
    Load all files associated with a task.
    Searches in both uploads/task_id/ and uploads/task_id/output/
    
    Args:
        task_id: Task ID
        
    Returns:
        List of FileHandler objects
    """
    # Get file paths from uploads directory
    upload_dir = Path("uploads") / task_id
    
    if not upload_dir.exists():
        return []
    
    files = []
    # Search in root directory
    for file_path in upload_dir.glob("*"):
        if file_path.is_file():
            try:
                files.append(handler)
            except Exception as e:
                pass

    # Also search in output subdirectory
    output_dir = upload_dir / "output"
    if output_dir.exists():
        for file_path in output_dir.glob("*"):
            if file_path.is_file():
                try:
                    files.append(handler)
                except Exception as e:
                    pass

    return files


async def save_output_file(
    task_id: str,
    filename: str,
    content: Union[str, bytes, list, dict],
    agent_id: str = None
) -> str:
    """
    Save output file for a task.
    Auto-converts data structures to appropriate format.
    
    Args:
        task_id: Task ID
        filename: Output filename
        content: File content (string, bytes, list of dicts, or dict)
        agent_id: Optional agent ID for logging
        
    Returns:
        File path
    """
    import json
    import csv
    import io
    
    # Auto-convert data to appropriate format
    if isinstance(content, list) and len(content) > 0 and isinstance(content[0], dict):
        # List of dicts -> CSV or Excel based on filename
        if filename.endswith('.xlsx'):
            # Convert to Excel format
            try:
                import openpyxl
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                
                # Write headers
                if content:
                    headers = list(content[0].keys())
                    ws.append(headers)
                    
                    # Write data rows
                    for row_dict in content:
                        row_values = [row_dict.get(header, '') for header in headers]
                        ws.append(row_values)
                
                # Save to bytes
                output = io.BytesIO()
                wb.save(output)
                content = output.getvalue()
                output.close()
            except ImportError:
                # Fallback to CSV if openpyxl not available
                filename = f"{Path(filename).stem}.csv"
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=content[0].keys())
                writer.writeheader()
                writer.writerows(content)
                content = output.getvalue()
        else:
            # Default to CSV
            if not filename.endswith('.csv'):
                filename = f"{Path(filename).stem}.csv"
            
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=content[0].keys())
            writer.writeheader()
            writer.writerows(content)
            content = output.getvalue()
    elif isinstance(content, str) and filename.endswith('.xlsx'):
        # CSV string -> Excel
        try:
            import openpyxl
            from openpyxl import Workbook
            
            # Parse CSV string
            csv_reader = csv.reader(io.StringIO(content))
            rows = list(csv_reader)
            
            wb = Workbook()
            ws = wb.active
            
            # Write all rows
            for row in rows:
                ws.append(row)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            content = output.getvalue()
            output.close()
        except ImportError:
            # Fallback: keep as CSV but change extension
            filename = f"{Path(filename).stem}.csv"
    elif isinstance(content, dict):
        # Dict -> JSON
        if not filename.endswith('.json'):
            filename = f"{Path(filename).stem}.json"
        content = json.dumps(content, indent=2, ensure_ascii=False)
    elif isinstance(content, (list, tuple)) and not isinstance(content, bytes):
        # List -> JSON
        if not filename.endswith('.json'):
            filename = f"{Path(filename).stem}.json"
        content = json.dumps(content, indent=2, ensure_ascii=False)
    
    # Create output directory
    output_dir = Path("uploads") / task_id / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    file_path = output_dir / filename
    
    # Write file
    if isinstance(content, bytes):
        with open(file_path, 'wb') as f:
            f.write(content)
    else:
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
            
    return str(file_path)

